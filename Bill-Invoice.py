from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

# Create a workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Invoice"

# Define styles
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# Invoice Header
ws["A1"] = "Invoice"
ws["A1"].font = Font(size=14, bold=True)
ws["A1"].alignment = center_align
ws.merge_cells("A1:D1")

# Invoice Metadata
ws["A3"] = "Customer Name:"
ws["B3"] = "John Doe"
ws["A4"] = "Invoice Date:"
ws["B4"] = datetime.now().strftime("%Y-%m-%d")

# Table Headers
ws["A6"] = "Item"
ws["B6"] = "Quantity"
ws["C6"] = "Price"
ws["D6"] = "Total"
for col in ["A", "B", "C", "D"]:
    ws[f"{col}6"].font = bold_font
    ws[f"{col}6"].border = border
    ws[f"{col}6"].alignment = center_align

# Sample Data
items = [("Widget A", 2, 50), ("Widget B", 1, 100), ("Widget C", 5, 20)]
row = 7
for item, qty, price in items:
    ws[f"A{row}"] = item
    ws[f"B{row}"] = qty
    ws[f"C{row}"] = price
    ws[f"D{row}"] = f"=B{row}*C{row}"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].border = border
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
    row += 1

# Total Row
ws[f"C{row}"] = "Total"
ws[f"C{row}"].font = bold_font
ws[f"C{row}"].alignment = center_align
ws[f"C{row}"].border = border
ws[f"D{row}"] = f"=SUM(D7:D{row-1})"
ws[f"D{row}"].font = bold_font
ws[f"D{row}"].alignment = center_align
ws[f"D{row}"].border = border

# Save the Invoice
wb.save("otP/Invoicegenerate.xlsx")
print("sucessfully created")