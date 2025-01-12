from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active

ws["A1"]="Hello Programmers! "

"formatting the cell dimensions"
ws["A1"].font= Font(bold=True , color="FF5733")
ws['A1'].alignment=Alignment(horizontal="center")

#merge cells
ws.merge_cells("A1:C1")

#Using Formuals
ws["A3"] = 10
ws["B3"] = 5
ws["C3"] = "=A3+B3"


wb.save("otP/5.working_with_cells.xlsx")
print("runned succesfully")